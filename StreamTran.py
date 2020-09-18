import numpy as np
from fipy import *
from pylab import *
from scipy.interpolate import interp1d
import pandas as pd
import datetime as dt
import pdb
import convolution_integral as ci
import lmfit as lm

def CreateLatFlowVec(ql,nx):
    ''' returns a vector of length nx (full grid - L/dx), with len(ql) evenly spaced steps, each step having the value ql_i.'''
    numSteps = len(ql)
    stepL = int(round(nx/numSteps))
    stepsI = arange(1,numSteps+1)*stepL
    q = np.zeros(int(nx)) 
    ict = 0
    for i in range(len(q)):
        step_i = stepsI[ict]
        if i >= step_i:
            ict += 1
        q[i]=q[i]+ql[ict]
    return q 

def CreateLmParams(x,name,lowerbound,upperbound):
    '''returns an lmfit paramters class with a parameter for each item in x (array like), each with a name (name_ict)'''
    ict = 0
    m = lm.Parameters()
    for p in x:
        name_i = name+'_'+str(ict)
        m.add(name_i,value=p,min=lowerbound,max=upperbound,vary=True)
        ict += 1
    return m

def AppendLmParams(m,x,name,lowerbound,upperbound):
    '''returns an lmfit paramters class (m) with a parameter added to m for each item in x (array like), each with a name (name_ict)'''
    ict = 0
    for p in x:
        name_i = name+'_'+str(ict)
        m.add(name_i,value=p,min=lowerbound,max=upperbound,vary=True)
        ict += 1
    return m

def ExtractLmParamVec(m,name):
    '''extracts an array x from all entries in params m with name 'name_i'. Expects the name to be split by an underscore.'''

    x = []  
    for k in sort(list(m.keys())):
        if k.split('_')[0] == name:
            x.append(m[k].value)
    x = np.array(x)
    return x
        

def InterpolateMatrialValue(x,V,nx,L):
    ''' linearly interpolate the set of distance (x), variable (V) measurements at each cell location.'''
    f = interp1d(x,V)
    dx = L/nx
    #ghetto way to do this.  Should figure out how get x defined from mesh.
    xnew = arange(0,L,dx)
    V_i = f(xnew)
    return V_i

def InterpolateTribPump(x,Q_t,nx,L):
    ''' Adds a pump discharge or Trib discharge for each pair of distance(x), discharge (Q_t) measurements.  Return a vector of length nx, with pump or trib discharge at the nearest mesh location.'''
    dx = L/nx
    xnew = arange(0,L,dx)
    Q = np.zeros(len(xnew))
    for i in range(len(x)):
        Q[np.argmin(abs(x[i]-xnew))]=Q_t[i]
    return Q

def InterpolateGwConc(x,C_gw,numSteps,nx,L):
    ''' Assigns concnetration for each groundwater inflow step (numSteps) given distance (x) and groundwater concentration (C_gw) pairs using nearest neighbor interpolation.  Returns the groundwater concentration for whole reach.'''
    # first nearest neighbor interpolation for all groundwater steps...
    nx = int(nx)
    stepL = int(round(L/numSteps))
    stepsI = arange(1,numSteps+1)*stepL
    stepsI[-1] = int(round(L))
    stepsM = stepsI/2.
    C_gw_s = np.zeros(numSteps)
    for i in range(numSteps):
        C_gw_s[i] = C_gw[np.argmin(abs(stepsM[i]-x))]

    #now the full nx vector
    Cgw = np.zeros(int(nx))
    xx = np.linspace(0,L,num=nx)
    ict = 0
    for i in range(int(nx)):
        xi = xx[i]
        step_i = stepsI[ict]
        if xi > step_i:
            ict += 1
        Cgw[i]=Cgw[i]+C_gw_s[ict]
    return Cgw 

def InterpolateGwConcTau(x,tau,numSteps,nx,L,C_t,tracer_t,t_i,lamba,mod_type='exponential'):
    ''' Assigns concnetration for each groundwater inflow step (numSteps) given distance (x) and groundwater mean age (tau) pairs using nearest neighbor interpolation.  Requires the precipitation input as a pandas time indexed data frame (C_t) and the tracer (tracer_t) for which to calculate the concentration, and the simulation time (t_i) as python datetime.  Returns the groundwater concentration for whole reach. Lambda needs to be days for convolution.  this is a different unit than StreamTran...'''
    
    #calculate groundwater concentrations
    C_gw = np.zeros(len(tau))
    for i in range(len(tau)):
        #C_t_c = ci.conv_int_fft(C_t,tracer_t,t_i,tau[i],lamba,mod_type=mod_type)
        #C_gw[i] = C_t_c[-1]
        #pdb.set_trace()
        C_t_c = ci.conv_int_discrete(C_t,tracer_t,t_i,tau[i],lamba,mod_type=mod_type)
        C_gw[i] = C_t_c
    
    Cgw = InterpolateGwConc(x,C_gw,numSteps,nx,L)

    return Cgw

class StreamTranSim(object):
    '''Simulation class for a stream transport simulation'''
    def __init__(self,name):
        self.name=name
        self.L = 0.  # reach length
        self.nx = 1.e4 # number of grid cells - default should be good for most applications
        self.Ai = ([0.],[0.]) # field x-sec area data tuple ([x-dist],[var]) for now
        self.wi = ([0.],[0.]) # field x-sec width data
        self.t_end = 100.*3.15e7 #end time
        self.nt = 100 #time steps - need to have a smart default here.
        self.q_lin = [0.] # groundwater input step funtion values for all steps, the lenth of this vector sets how many steps there will be
        self.q_lo = [0.] # groundwater loss step funtion values for all steps
        self.Q_us = 1. #upstream discharge
        self.P = 0. #precipitation
        self.E = 0. #evaporation
        self.Q_trib = ([0.],[0.],[0.]) #location, discharge, concentration for tributaries
        self.Pump = ([0.],[0.]) #location discharge for pumping
        self.C_t = pd.DataFrame() #historical preciptation input for tracers
        self.SimRes = SimRes()
        self.Tracers = {}
        self.mesh = Grid1D(nx = self.nx, dx = self.L/self.nx)
        self.tau_flag = 0 #are groundwater concentrations to be calculated from mean age
        self.C_tau = ([0.],[0.]) #groundwater age should be independent of tracer concentration
        self.age_dist = 'exponential'
        self.t_i = [] #default the simulation date to the end of the concentration input data series.

    def CalcFlow(self):
        '''Calculate the river discharge given the tracer simulation class (TS)'''
        #set up simulation
        #discretization and interpolation
        mesh = Grid1D(nx = self.nx, dx = self.L/self.nx)
        self.mesh = mesh
        A_i = InterpolateMatrialValue(self.Ai[0],self.Ai[1],self.nx,self.L)
        w_i = InterpolateMatrialValue(self.wi[0],self.wi[1],self.nx,self.L)
        d_i = A_i/w_i
        # q_lin, q_lo
        P = self.P
        E = self.E
        self.SimRes.x = mesh.cellCenters.value[0,:]
        Q_trib = InterpolateTribPump(self.Q_trib[0],self.Q_trib[1],self.nx,self.L)
        Pump = InterpolateTribPump(self.Pump[0],self.Pump[1],self.nx,self.L)
        self.SimRes.Q_trib = CellVariable(name="Tributary Flow",mesh=self.mesh,value=Q_trib)
        self.SimRes.Pump = CellVariable(name="Pump",mesh=self.mesh,value=Pump)
        self.SimRes.q_lin = CreateLatFlowVec(self.q_lin,self.nx)
        self.SimRes.q_lo = CreateLatFlowVec(self.q_lo,self.nx)
        
        #matrial property cell variables
        self.SimRes.A = CellVariable(name="StreamXCArea",mesh=self.mesh,value=A_i)
        self.SimRes.d = CellVariable(name="StreamDepth",mesh=self.mesh,value=d_i)
        self.SimRes.w = CellVariable(name="StreamWidth",mesh=self.mesh,value=w_i)
        #think gw discharges and pump/trib discharge should be cell variables...
        self.SimRes.q_lin = CellVariable(name="GW_Discharge",mesh=self.mesh,value=self.SimRes.q_lin)
        self.SimRes.q_lo = CellVariable(name="GW_Recharge",mesh=self.mesh,value=self.SimRes.q_lo)

        #dependent cell variables
        Q = CellVariable(name="Discharge",mesh=mesh,value=self.Q_us)
        Q.constrain(self.Q_us, mesh.facesLeft)
        Q.faceGrad.constrain(0.,mesh.facesRight)
        eqD = PowerLawConvectionTerm(coeff=(1,),var=Q) ==  P*self.SimRes.w - E*self.SimRes.w + self.SimRes.q_lin*self.SimRes.w - self.SimRes.q_lo*self.SimRes.w + self.SimRes.Q_trib/self.mesh.dx - self.SimRes.Pump
        eqD.solve(var=Q)
        self.SimRes.Q = Q
    
    def add_tracer(self,name):
        '''add a tracer of name to the simulation'''
        self.Tracers[name]=Tracer(name)

    def CalcTran(self):
        '''Calculate the tracer transport concentration'''
        self.CalcFlow()
        k = iter(self.Tracers.keys())
        for kk in k:
            #add results classes
            self.SimRes.add_tracer(kk)
            #dependent cell variables
            #C = CellVariable(name="Concentration",mesh=self.mesh,value=self.Tracers[kk].C_us,hasOld=1)
            C = CellVariable(name="Concentration",mesh=self.mesh,value=self.Tracers[kk].C_us) #steady state
            #interpolate the groundwater concentration...
            if self.tau_flag:
                if self.Tracers[kk].has_age:
                    Cgwi = InterpolateGwConcTau(self.C_tau[0],self.C_tau[1],len(self.C_tau[0]),self.nx,self.L,self.C_t,kk,self.t_i,self.Tracers[kk].lamma,mod_type=self.age_dist)
                    Ctaui = InterpolateGwConc(self.C_tau[0],self.C_tau[1],len(self.C_tau[0]),self.nx,self.L)
                    self.SimRes.Ctaui = Ctaui
                else:
                    Cgwi = InterpolateGwConc(self.Tracers[kk].C_gw[0],self.Tracers[kk].C_gw[1],len(self.q_lin),self.nx,self.L)
    
            else:
                Cgwi = InterpolateGwConc(self.Tracers[kk].C_gw[0],self.Tracers[kk].C_gw[1],len(self.q_lin),self.nx,self.L)
        
            Cgwi = CellVariable(name="GW Concentration",mesh=self.mesh,value=Cgwi)
            C_trib = InterpolateTribPump(self.Q_trib[0],self.Q_trib[2][kk],self.nx,self.L)
            C_trib = CellVariable(name="Trib Concentration",mesh=self.mesh,value=C_trib)
            self.SimRes.tracers[kk].C_trib = C_trib
            self.SimRes.tracers[kk].Cgwi=Cgwi

            #boundary conditions
            C.constrain(self.Tracers[kk].C_us, self.mesh.facesLeft)
            C.faceGrad.constrain(0., self.mesh.facesRight)
            #setup equations
#            if disp_flag:
#                eqT = PowerLawConvectionTerm((1.,),var=C) == DiffusionTerm(coeff=self.Tracers[kk].D*self.SimRes.w*self.SimRes.d/self.SimRes.Q, var=C) + self.SimRes.w*self.SimRes.q_lin/self.SimRes.Q*self.SimRes.tracers[kk].Cgwi - ImplicitSourceTerm(coeff=self.SimRes.w*self.SimRes.q_lin/self.SimRes.Q,var=C) - ImplicitSourceTerm(coeff=self.SimRes.w*self.Tracers[kk].k/self.SimRes.Q,var=C) + self.SimRes.w*self.Tracers[kk].k/self.SimRes.Q*self.Tracers[kk].C_atm + self.SimRes.Q_trib/self.SimRes.Q/self.mesh.dx*self.SimRes.tracers[kk].C_trib - ImplicitSourceTerm(coeff=self.SimRes.Q_trib/self.SimRes.Q/self.mesh.dx,var=C) - ImplicitSourceTerm(coeff=self.SimRes.w*self.SimRes.d/self.SimRes.Q*self.Tracers[kk].lamma,var=C)
#            else:
            eqT = PowerLawConvectionTerm((1.,),var=C) == self.SimRes.w*self.SimRes.q_lin/self.SimRes.Q*self.SimRes.tracers[kk].Cgwi - ImplicitSourceTerm(coeff=self.SimRes.w*self.SimRes.q_lin/self.SimRes.Q,var=C) - ImplicitSourceTerm(coeff=self.SimRes.w*self.Tracers[kk].k/self.SimRes.Q,var=C) + self.SimRes.w*self.Tracers[kk].k/self.SimRes.Q*self.Tracers[kk].C_atm + self.SimRes.Q_trib/self.SimRes.Q/self.mesh.dx*self.SimRes.tracers[kk].C_trib - ImplicitSourceTerm(coeff=self.SimRes.Q_trib/self.SimRes.Q/self.mesh.dx,var=C) - ImplicitSourceTerm(coeff=self.SimRes.w*self.SimRes.d/self.SimRes.Q*self.Tracers[kk].lamma,var=C) + ImplicitSourceTerm(coeff=self.E*self.SimRes.w,var=C) - ImplicitSourceTerm(coeff=self.P*self.SimRes.w,var=C)#need to add evaporation and precip.
            eqT.solve(var=C)
            self.SimRes.tracers[kk].C = C

    def MisfitFlux(self,m,tracers,FieldData,DischData,use_disch=1.,DischargeError=0.10,log_flag=1):
        '''Calculate the misfit between modeled and the observed data for the list of tracers in tracers.  FieldData is a dictionary keyed by the tracer name with 'tracer':(xd,yd). Returns the redisidual for a least squares inversion.'''
        self.q_lin = ExtractLmParamVec(m,'qlin')
        if log_flag:
            self.q_lin = np.exp(self.q_lin)
        mm = len(self.q_lin)
        print('Current Params:')
        print(self.q_lin)
        self.CalcTran()
        xd = DischData[0]
        yd = DischData[1]
        xm = self.mesh.cellCenters()[0,:]
        xm[0] = 0.
        xm[-1] = self.L
        ym = self.SimRes.Q.value
        #calculatate the misfit given the current groundwater flux
        if use_disch:
            f = interp1d(xm,ym,kind='linear')
            yms = f(xd)
            self.SimRes.resid = (yd-yms)/(DischargeError*np.mean(yd))
        for tracer in tracers:
            xd = FieldData[tracer][0]
            yd = FieldData[tracer][1]
            ym = self.SimRes.tracers[tracer].C.value
            f = interp1d(xm,ym,kind='linear')
            yms = f(xd)
            self.SimRes.resid = np.append(self.SimRes.resid,(yd-yms)/(self.Tracers[tracer].error_perc*np.mean(yd)))
            #self.SimRes.resid = np.append(self.SimRes.resid,yd-yms)
        nn = len(self.SimRes.resid)
        if mm>nn:
            print('number of estimated groundwater inflow steps greater than constraining data, reduce number of steps.')    
        print('residual = ' + str(np.linalg.norm(self.SimRes.resid)))
        return self.SimRes.resid

    def MisfitAge(self,m,tracers,FieldData,DischData,use_disch=0,DischargeError=0.10):
        '''Calculate the misfit between modeled and the observed data for the list of tracers in tracers with the flag has_age.  FieldData is a dictionary keyed by the tracer name with 'tracer':(xd,yd). User has option to fit discharge or not with the use_disch flag. Returns the redisidual for a least squares inversion.'''
        C_tau = ExtractLmParamVec(m,'cTau')
        self.C_tau = (self.C_tau[0],C_tau)
        for tracer in tracers:
            if self.Tracers[tracer].has_age:
                self.Tracers[tracer].C_gw=self.C_tau
        print('Current Groundwater Age:')
        print(self.C_tau[1]/365.)
        self.CalcTran()
        xm = self.mesh.cellCenters()[0,:]
        xm[0] = 0.
        xm[-1] = self.L
        if use_disch:
            xd = DischData[0]
            yd = DischData[1]
            ym = self.SimRes.Q.value
            f = interp1d(xm,ym,kind='linear')
            yms = f(xd)
            self.SimRes.resid = (yd-yms)/(DischargeError*np.mean(yd))
        else:
            self.SimRes.resid = np.array([])
        #self.SimRes.resid = yd-yms
        #calculatate the misfit given the current groundwater flux
        for tracer in tracers:
            if self.Tracers[tracer].has_age:
                xd = FieldData[tracer][0]
                yd = FieldData[tracer][1]
                ym = self.SimRes.tracers[tracer].C.value
                f = interp1d(xm,ym,kind='linear')
                yms = f(xd)
                self.SimRes.resid = np.append(self.SimRes.resid,(yd-yms)/(self.Tracers[tracer].error_perc*np.mean(yd)))
            #self.SimRes.resid = np.append(self.SimRes.resid,yd-yms)
            
        print('residual = ' + str(np.linalg.norm(self.SimRes.resid)))
        return self.SimRes.resid

    def MisfitGWConc(self,m,tracers,FieldData,DischData,use_disch=0,DischargeError=0.10,log_flag_c=1):
        '''Calculate the misfit between modeled and the observed data for the list of tracers in tracers with the flag has_age.  FieldData is a dictionary keyed by the tracer name with 'tracer':(xd,yd). User has option to fit discharge or not with the use_disch flag. Returns the redisidual for a least squares inversion.'''
        for tracer in tracers:
            C_gw = ExtractLmParamVec(m,'Cgw'+tracer)
            if log_flag_c:
                C_gw = np.exp(C_gw)
                
            self.Tracers[tracer].C_gw = (self.Tracers[tracer].C_gw[0],C_gw)
        print('Current Params:')
        for tracer in tracers:
            print(tracer)
            print(self.Tracers[tracer].C_gw)
        self.CalcTran()
        xm = self.mesh.cellCenters()[0,:]
        xm[0] = 0.
        xm[-1] = self.L
        if use_disch:
            xd = DischData[0]
            yd = DischData[1]
            ym = self.SimRes.Q.value
            f = interp1d(xm,ym,kind='linear')
            yms = f(xd)
            self.SimRes.resid = (yd-yms)/(DischargeError*np.mean(yd))
        else:
            self.SimRes.resid = np.array([])
        #self.SimRes.resid = yd-yms
        #calculatate the misfit given the current groundwater flux
        for tracer in tracers:
            xd = FieldData[tracer][0]
            yd = FieldData[tracer][1]
            ym = self.SimRes.tracers[tracer].C.value
            f = interp1d(xm,ym,kind='linear')
            yms = f(xd)
            self.SimRes.resid = np.append(self.SimRes.resid,(yd-yms)/(self.Tracers[tracer].error_perc*np.mean(yd)))
            #self.SimRes.resid = np.append(self.SimRes.resid,yd-yms)
            
        print('residual = ' + str(np.linalg.norm(self.SimRes.resid)))
        return self.SimRes.resid

    def CalcGwInflow(self,tracers,FieldData,DischData,DischargeError=0.10,use_disch=1,log_flag=1):
        '''Estimate the best groundwater inflow to match the given stream concentrations.'''
        qlin = self.q_lin
        lb = 0.
        ub = 1.
        if log_flag:
            qlin = np.log(qlin)
            lb = -20
            up = 0.
        m = CreateLmParams(qlin,'qlin',lb,ub)
        mini = lm.Minimizer(self.MisfitFlux,m,fcn_args=(tracers,FieldData,DischData),fcn_kws={'DischargeError':DischargeError,'use_disch':use_disch,'log_flag':log_flag})
        fit = mini.minimize(params=m,ftol=1.e-5,gtol=1.e-5,xtol=1e-5)
        print(lm.fit_report(fit))
        self.q_lin = ExtractLmParamVec(fit.params,'qlin')
        if log_flag:
            self.q_lin = np.exp(self.q_lin)
        self.SimRes.fit_solution = fit
        #run a last best solution
        self.CalcTran()

    def CalcGwConc(self,tracers,FieldData,DischData,use_disch=0,DischargeError=0.10,log_flag_c=1):
        '''Estimate the best groundwater concentration to match the given stream concentrations of tracers '''
        m = lm.Parameters()
        for tracer in tracers:
            C_gw = self.Tracers[tracer].C_gw[1] #the user inputs the number of C_gw steps on model initialization... The number of steps should not exceed the number of samples for inversion.
            lb = self.Tracers[tracer].C_gw_lb
            ub = self.Tracers[tracer].C_gw_ub
            if log_flag_c:
                C_gw = np.log(C_gw)
                lb = np.log(lb)
                ub = np.log(ub)
            m = AppendLmParams(m,C_gw,'Cgw'+tracer,lb,ub)

        mini = lm.Minimizer(self.MisfitGWConc,m,fcn_args=(tracers,FieldData,DischData),fcn_kws={'use_disch':use_disch,'DischargeError':DischargeError,'log_flag_c':log_flag_c})
        fit = mini.minimize(params=m,xtol=1.e-5,ftol=1e-5,epsfcn=1.e-3)
#        fit = mini.minimize(params=m)
        for tracer in tracers:
            C_gw = ExtractLmParamVec(fit.params,'Cgw'+tracer)
            if log_flag_c:
                C_gw = np.exp(C_gw)
            self.Tracers[tracer].C_gw = (self.Tracers[tracer].C_gw[0],C_gw)
        self.SimRes.fit_solution = fit
        self.CalcTran()
        print(lm.fit_report(fit))
        #run a last best solution

    def CalcGwAge(self,tracers,FieldData,DischData,use_disch=0,DischargeError=0.10):
        '''Estimate the best groundwater mean age to match the given stream concentrations of tracers with age information - has_age=1.'''
        C_tau = self.C_tau[1]
        lb = 1.
        ub = 1000.*365.  #upper limit of 100 years mean age right now.  Making this longer will make inherent problems with the convolution technique be a problem. Needs to be dealt with at some point
        m = CreateLmParams(C_tau,'cTau',lb,ub)
        mini = lm.Minimizer(self.MisfitAge,m,fcn_args=(tracers,FieldData,DischData),fcn_kws={'use_disch':use_disch,'DischargeError':DischargeError})
#        fit = mini.minimize(params=m,xtol=1.e-5,ftol=1e-5,epsfcn=1.e-3)
        fit = mini.minimize(params=m)
        C_tau = ExtractLmParamVec(fit.params,'cTau')
        self.C_tau = (self.C_tau[0],C_tau)
        self.SimRes.fit_solution = fit
        self.CalcTran()
        print(lm.fit_report(fit))
        #run a last best solution

class SimRes(object):
    '''Simulation Results.'''
    def __init__(self):
        self.x = [] # grid cell centers
        self.Q = [] # modeled discharge
        self.A = []
        self.d = []
        self.w = []
        self.q_lin = []
        self.q_lo = []
        self.Q_trib = []
        self.Pump = []
        self.tracers = {} #dictionary of TracerRes objects
        self.Ctaui = []

    def add_tracer(self,name):
        self.tracers[name]=TracerRes(name)

    def dump2xl(self,wb_name="model_results.xlsx"):
        '''Dump the SimRes interpolated model results to an excel file'''
        from openpyxl import Workbook
        
        wb = Workbook()
        dest_file = wb_name
        ws1 = wb.active
        ws1.title = 'model results'
        output_vars = ['x','Q','d','w','q_lin','Ctaui','Q_trib']
        output_var_notes = ['Distance Downstream','Modeled Discharge (l^3/t)','Interpolated Depth (l)','Interpolated Width (l)','Groundwater Lateral flux (l/t)','Estmated Mean Groundwater Age (days)','Tributary Discharge (l^3/t)']
        li = len(output_vars)+1
        
        for i in range(1,li):
            j=1
            ws1.cell(row=j,column=i,value=output_vars[i-1])
            j=2 
            ws1.cell(row=j,column=i,value=output_var_notes[i-1])
            for j in range(3,len(self.x)+3):
                try:
                    ws1.cell(row=j,column=i,value=self.__dict__[output_vars[i-1]].value[j-3])
                except AttributeError:
                    ws1.cell(row=j,column=i,value=self.__dict__[output_vars[i-1]][j-3])
                    
        tk = list(self.tracers.keys())
        for i in range(li,len(tk)+li):
            j=1
            ws1.cell(row=j,column=i,value=tk[i-li])
            j=2 
            ws1.cell(row=j,column=i,value='modeled stream conc.')
            for j in range(3,len(self.x)+3):
                ws1.cell(row=j,column=i,value=self.tracers[tk[i-li]].C.value[j-3])
                
        wb.save(filename = dest_file)         

            

class TracerRes(object):
    '''Modeled Tracer profiles'''
    def __init__(self,name):
        self.name = name
        self.C = []
        self.Cgwi = []
        self.C_trib = []

class Tracer(object):
    '''Tracer specific data'''
    def __init__(self,name):
        self.name = name
        self.C_us = 1. #upstream concentration - constant for now, but should be a time, concentration data set.
        self.C_gw = ([0.],[0.]) # ground concentration or mean age, data tuple ([x-dist],[var]) - how it gets interpolated depends upon the tau_flag attribute of the Simulation class.
        self.C_gw_ub = 0. #the upper and lower bounds of concentration in groundwater for inversion
        self.C_gw_lb = 1000.
        self.D = 6.e-9 # Dispersion coefficient - default is diffusion only
        self.k = 2./60./60./24. # Gas exchange coefficient
        self.lamma = 0. # decay coefficient for tracer
        self.C_atm = 0. # equilibrium atmospheric concentration
        self.error_perc = 0.05
        self.has_age = 0.
        self.unit = ''


